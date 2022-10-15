import openpyxl

file_path = 'desktop/tgt/input.xlsx'
#file_path2 = 'desktop/tgt/output.xlsx'
week_days = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота', 'Воскресенье']
wk_days = {'Пн':0, 'Вт':1, 'Ср':2, 'Чт':3, "Пт":4, "Сб" : 5, "Вс" : 6}


def yaxlitlash(time):
    st = time.split(':')
    if len(st) > 1:
        if int(st[1]) >= 30:
            time = str(int(st[0].strip())+1)
        else:
            time = str(int(st[0].strip()))
    return time


def parse_data(data, code, week_days = week_days, h = []):
    date1 = data.split('; ')
    date = data.split(': ')
    if date[0] == 'Ежедневно':
        df = data.split(" (обед: ")
        dam_st_time, dam_ed_time = 0, 0
        if len(df) > 1:
            dat_l = df[0]
            date = dat_l.split(': ')
            dam_times = df[1].strip().split('-')
            dam_st_time = yaxlitlash(dam_times[0])
            dam_ed_time = yaxlitlash(dam_times[1][:-1])
        time = str(date[1]).strip().split('-')
        st_time, ed_time = time[0], time[1]
        st = yaxlitlash(st_time)
        ed = yaxlitlash(ed_time)
        for day in week_days:
            h.append([code, st, ed, dam_st_time, dam_ed_time, day])
    elif date[0] == 'Круглосуточно':
        for day in week_days:
            h.append([code, '0', '24', '0', '0', day])
    elif len(date1) > 1:
        z = []
        for i in date1:
            i = i.strip()
            data_f = i.split(': ')
            wk_days_time = data_f[0].split('-')


            if data_f[1] == 'Круглосуточно':
                st = '0'
                ed = '24'
                dam_st_time = '0'
                dam_ed_time = '0'
                if len(wk_days_time) > 1:
                    st_wk_time = wk_days[str(wk_days_time[0]).strip()]
                    ed_wk_time = wk_days[str(wk_days_time[1]).strip()]
                    for i in range(st_wk_time, ed_wk_time+1):
                        h.append([code, st, ed, dam_st_time, dam_ed_time, week_days[i]])
                else:
                    wk_day = wk_days[str(wk_days_time[0]).strip()]
                    h.append([code, st, ed, dam_st_time, dam_ed_time, week_days[wk_day]])
            elif data_f[1] == 'Выходной':
                st = '0'
                ed = '0'
                dam_st_time = '0'
                dam_ed_time = '24'
                if len(wk_days_time) > 1:
                    st_wk_time = wk_days[str(wk_days_time[0]).strip()]
                    ed_wk_time = wk_days[str(wk_days_time[1]).strip()]
                    for i in range(st_wk_time, ed_wk_time+1):
                        h.append([code, st, ed, dam_st_time, dam_ed_time, week_days[i]])
                else:
                    wk_day = wk_days[str(wk_days_time[0]).strip()]
                    h.append([code, st, ed, dam_st_time, dam_ed_time, week_days[wk_day]])
            else:
                if len(data_f) > 2:
                    dam_times = str(data_f[-1])[:-1].strip().split('-')
                    dam_st_time = yaxlitlash(dam_times[0])
                    dam_ed_time = yaxlitlash(dam_times[1][:-1])
                else:
                    dam_st_time, dam_ed_time = 0, 0
                days_time_2 = data_f[1].split(' (')[0].strip().split('-')
                st_time, ed_time = days_time_2[0], days_time_2[1]
                st = yaxlitlash(st_time)
                ed = yaxlitlash(ed_time)

            
            if len(wk_days_time) > 1:
                st_wk_time = wk_days[str(wk_days_time[0]).strip()]
                ed_wk_time = wk_days[str(wk_days_time[1]).strip()]
                for i in range(st_wk_time, ed_wk_time+1):
                    h.append([code, st, ed, dam_st_time, dam_ed_time, week_days[i]])
            else:
                wk_day = wk_days[str(wk_days_time[0]).strip()]
                h.append([code, st, ed, dam_st_time, dam_ed_time, week_days[wk_day]])
    return h  
            #z.append(data_f)
        #print(z)
    



def edit_file(path = file_path):
    workbook = openpyxl.load_workbook(path)
    workbook2 = openpyxl.Workbook()
    sheet2 = workbook2.active
    sheet = workbook.active
    all_of_all = []
    sheet2.column_dimensions['A'].width = 20.0
    sheet2.column_dimensions['B'].width = 8.0
    sheet2.column_dimensions['C'].width = 8.0
    sheet2.column_dimensions['D'].width = 8.0
    sheet2.column_dimensions['E'].width = 8.0
    sheet2.column_dimensions['F'].width = 12.0
    for i in range(2, sheet.max_row+1):  #sheet.max_row+1
        code, data = sheet.cell(i, 1).value, sheet.cell(i, 2).value
        necessary = parse_data(data, code)
        all_of_all.extend(necessary)
    for all in range(len(all_of_all)):
        code, st, ed, dam_st, dam_ed, day = all_of_all[all][0], all_of_all[all][1], all_of_all[all][2], all_of_all[all][3], all_of_all[all][4], all_of_all[all][5]
        sheet2.cell(all+2, 1).value, sheet2.cell(all+2, 2).value, sheet2.cell(all+2, 3).value, sheet2.cell(all+2, 4).value, sheet2.cell(all+2, 5).value, sheet2.cell(all+2, 6).value = code, int(st), int(ed), int(dam_st), int(dam_ed), day
    workbook2.save('desktop/tgt/output1.xlsx')


edit_file()